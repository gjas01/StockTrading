from __future__ import annotations

import tempfile
import tkinter as tk
import webbrowser
from html import escape
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from src import db
from src.services.livermore_ledger import LedgerRow, build_ledger, threshold_labels


COLUMN_HEADERS = {
    "secondary_rally":    "Secondary Rally",
    "natural_rally":      "Natural Rally",
    "upward_trend":       "Upward Trend",
    "downward_trend":     "Downward Trend",
    "natural_reaction":   "Natural Reaction",
    "secondary_reaction": "Secondary Reaction",
}

# Full-width columns used in the 2-panel HTML/legacy view
TREE_COLUMNS = [
    ("date",               "Date",              92, "w"),
    ("high",               "High",              72, "e"),
    ("low",                "Low",               72, "e"),
    ("close",              "Close",             72, "e"),
    ("secondary_rally",    "Secondary Rally",   96, "e"),
    ("natural_rally",      "Natural Rally",     96, "e"),
    ("upward_trend",       "Upward Trend",      96, "e"),
    ("downward_trend",     "Downward Trend",    96, "e"),
    ("natural_reaction",   "Natural Reaction", 108, "e"),
    ("secondary_reaction", "Secondary Reaction",112,"e"),
    ("note",               "Note",             220, "w"),
]

# Compact columns for the 3-panel on-screen view (abbreviated headings, tighter widths)
TREE_COLUMNS_COMPACT = [
    ("date",               "Date",    76, "w"),
    ("high",               "High",    58, "e"),
    ("low",                "Low",     58, "e"),
    ("close",              "Close",   58, "e"),
    ("secondary_rally",    "Sec Rly", 64, "e"),
    ("natural_rally",      "Nat Rly", 64, "e"),
    ("upward_trend",       "Up Trnd", 64, "e"),
    ("downward_trend",     "Dn Trnd", 64, "e"),
    ("natural_reaction",   "Nat Rct", 68, "e"),
    ("secondary_reaction", "Sec Rct", 68, "e"),
    ("note",               "Note",   130, "w"),
]


# ---------------------------------------------------------------------------
# Cell formatting helpers
# ---------------------------------------------------------------------------

def _format_cell(row: LedgerRow, key: str) -> tuple[str, bool, bool, bool]:
    value = getattr(row, key)
    if value is None:
        return "", False, False, False
    return (
        f"{value:.2f}",
        key in row.pivotal,
        key in row.blue,
        key in row.red,
    )


def _display_cell(row: LedgerRow, key: str) -> str:
    value, pivotal, blue, red = _format_cell(row, key)
    if not value:
        return ""
    markers = ""
    if pivotal:
        markers += "*"
    if blue:
        markers += "^"
    if red:
        markers += "!"
    return f"{value}{markers}"


def _row_values(row: LedgerRow) -> tuple[str, ...]:
    values = [
        row.trade_date.isoformat(),
        f"{row.high:.2f}",
        f"{row.low:.2f}",
        f"{row.close:.2f}",
    ]
    for key in COLUMN_HEADERS:
        values.append(_display_cell(row, key))
    values.append(row.note)
    return tuple(values)


# ---------------------------------------------------------------------------
# HTML generation
# ---------------------------------------------------------------------------

def _ledger_table_html(title: str, rows: list[LedgerRow], price_count: int) -> str:
    if not rows:
        return f"<h2>{escape(title)}</h2><p>No price history available.</p>"

    header = (
        "<tr>"
        "<th>Date</th><th>High</th><th>Low</th><th>Close</th>"
        + "".join(f"<th>{escape(COLUMN_HEADERS[key])}</th>" for key in COLUMN_HEADERS)
        + "<th>Note</th></tr>"
    )
    body_rows = []
    for row in rows:
        cells = [
            f"<td>{row.trade_date.isoformat()}</td>",
            f"<td>{row.high:.2f}</td>",
            f"<td>{row.low:.2f}</td>",
            f"<td>{row.close:.2f}</td>",
        ]
        for key in COLUMN_HEADERS:
            text, pivotal, blue, red = _format_cell(row, key)
            css_classes = []
            if pivotal:
                css_classes.append("pivot")
            if blue:
                css_classes.append("near-upper")
            if red:
                css_classes.append("near-lower")
            css = f' class="{" ".join(css_classes)}"' if css_classes else ""
            cells.append(f"<td{css}>{escape(text)}</td>")
        cells.append(f"<td>{escape(row.note)}</td>")
        body_rows.append("<tr>" + "".join(cells) + "</tr>")

    subtitle = (
        f"{price_count} trading day(s). Underline = pivot point. "
        f"Blue = within 1.5% of upper pivot; red = within 1.5% of lower pivot."
    )
    return f"""
    <section class="ledger-block">
      <h2>{escape(title)}</h2>
      <p class="subtitle">{escape(subtitle)}</p>
      <table>
        <thead>{header}</thead>
        <tbody>{"".join(body_rows)}</tbody>
      </table>
    </section>
    """


def build_group_ledger_html(
    group_title: str,
    stocks: list[tuple[str, list[LedgerRow], int, str, str]],
) -> str:
    """Build a 3-column HTML page for a stock group.

    ``stocks`` is a list of (title, rows, price_count, reversal_label,
    continuation_label) tuples — one entry per stock.
    """
    thresholds_note = " / ".join(
        f"{title.split(':')[0]}: {rev}/{cont}"
        for title, _, _, rev, cont in stocks
    )
    table_html = "\n".join(
        _ledger_table_html(title, rows, count)
        for title, rows, count, _rev, _cont in stocks
    )
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>{escape(group_title)} - Livermore Ledgers</title>
  <style>
    body {{
      font-family: "Segoe UI", Arial, sans-serif;
      margin: 24px;
      color: #111;
    }}
    h1 {{ margin-bottom: 4px; }}
    .intro {{ margin-bottom: 24px; color: #333; font-size: 13px; }}
    .ledgers {{
      display: grid;
      grid-template-columns: 1fr 1fr 1fr;
      gap: 20px;
    }}
    .ledger-block {{ break-inside: avoid; }}
    .subtitle {{ font-size: 11px; color: #444; margin-top: 0; }}
    table {{
      border-collapse: collapse;
      width: 100%;
      font-size: 10px;
    }}
    th, td {{
      border: 1px solid #999;
      padding: 3px 5px;
      text-align: right;
    }}
    th:first-child, td:first-child,
    th:last-child,  td:last-child {{
      text-align: left;
    }}
    th {{ background: #eee; }}
    .pivot {{ text-decoration: underline; font-weight: 600; }}
    .near-upper {{ color: #0b5cab; font-weight: 600; }}
    .near-lower {{ color: #b00020; font-weight: 600; }}
    td.near-upper.pivot, td.near-lower.pivot {{ text-decoration: underline; }}
    @media print {{
      body {{ margin: 8px; font-size: 9px; }}
      .ledgers {{ grid-template-columns: 1fr 1fr 1fr; }}
      .no-print {{ display: none; }}
    }}
  </style>
</head>
<body>
  <h1>{escape(group_title)}</h1>
  <p class="intro">
    Jesse Livermore Market Key ledger. Trend established after initial reversal threshold
    from first close. Blank cells = no new extreme. Underline = confirmed pivot.
    Blue / red = within 1.5% of upper / lower pivot.<br>
    Thresholds (reversal / continuation): {escape(thresholds_note)}
  </p>
  <div class="ledgers">
    {table_html}
  </div>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _export_to_excel(
    group: dict,
    stocks: list[tuple[str, list[LedgerRow], int, tuple[str, str, str]]],
) -> None:
    """Save a workbook with one sheet per stock to a user-chosen path."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
    except ImportError:
        messagebox.showerror(
            "Missing dependency",
            "openpyxl is required for Excel export.\n"
            "Install it with:  pip install openpyxl",
        )
        return

    group_name = group.get("GroupName", "group")
    path = filedialog.asksaveasfilename(
        title="Export Ledgers to Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
        initialfile=f"ledger_{group_name.replace(' ', '_')}.xlsx",
    )
    if not path:
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the default blank sheet

    header_font   = Font(bold=True)
    header_fill   = PatternFill("solid", fgColor="DDDDDD")
    thin_side     = Side(style="thin", color="999999")
    cell_border   = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    center_align  = Alignment(horizontal="center")
    right_align   = Alignment(horizontal="right")

    col_headers = (
        ["Date", "High", "Low", "Close"]
        + list(COLUMN_HEADERS.values())
        + ["Note"]
    )
    col_keys = (
        ["date", "high", "low", "close"]
        + list(COLUMN_HEADERS.keys())
        + ["note"]
    )
    # Column widths (characters)
    col_widths = [12, 10, 10, 10, 14, 14, 14, 14, 16, 16, 40]

    for title, rows, _price_count, _thresholds in stocks:
        sheet_name = title.split(":")[0].strip()[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        for col_idx, (header, width) in enumerate(zip(col_headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.border = cell_border
            cell.alignment = center_align
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = width

        ws.freeze_panes = "A2"

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            # Date
            ws.cell(row=row_idx, column=1, value=row.trade_date.isoformat()).border = cell_border

            # OHLC numeric values
            for col_idx, attr in enumerate(["high", "low", "close"], 2):
                c = ws.cell(row=row_idx, column=col_idx, value=getattr(row, attr))
                c.number_format = "0.00"
                c.alignment     = right_align
                c.border        = cell_border

            # Six Livermore ledger columns (cols 5–10)
            for key_offset, key in enumerate(COLUMN_HEADERS.keys()):
                col_idx = 5 + key_offset
                text, pivotal, is_blue, is_red = _format_cell(row, key)
                c = ws.cell(row=row_idx, column=col_idx)
                c.border    = cell_border
                c.alignment = right_align
                if text:
                    c.value = float(text)
                    c.number_format = "0.00"
                    font_kwargs: dict = {}
                    if pivotal:
                        font_kwargs["bold"]      = True
                        font_kwargs["underline"] = "single"
                    if is_blue:
                        font_kwargs["color"] = "0B5CAB"
                    if is_red:
                        font_kwargs["color"] = "B00020"
                    if font_kwargs:
                        c.font = Font(**font_kwargs)

            # Note (last column)
            note_col = len(col_headers)
            c = ws.cell(row=row_idx, column=note_col, value=row.note)
            c.border    = cell_border
            c.alignment = Alignment(horizontal="left", wrap_text=True)

        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(col_headers)).coordinate}"

    wb.save(path)
    messagebox.showinfo("Export complete", f"Ledgers saved to:\n{path}")


# ---------------------------------------------------------------------------
# Ledger panel widget
# ---------------------------------------------------------------------------

class _LedgerPanelWidgets:
    def __init__(self, frame, title_label, subtitle_label, tree):
        self.frame    = frame
        self.title    = title_label
        self.subtitle = subtitle_label
        self.tree     = tree


# ---------------------------------------------------------------------------
# Main ledger sheet window
# ---------------------------------------------------------------------------

class LedgerSheetWindow(ttk.Frame):
    """Separate printable ledger view for a stock group (3 stocks)."""

    def __init__(self, master, group: dict):
        super().__init__(master, padding=12)
        self.group = group
        self._html = ""
        # Populated in load_ledgers: list of (title, rows, count, thresholds)
        self._stocks_data: list[tuple[str, list[LedgerRow], int, tuple[str, str, str]]] = []

        group_name = group.get("GroupName", f"Group #{group['GroupID']}")
        tickers = (
            f"{group['Stock1Ticker']} / {group['Stock2Ticker']} / {group['Stock3Ticker']}"
        )
        window_title = f"{group_name}  ({tickers})"

        # ── Toolbar ───────────────────────────────────────────────────────
        toolbar = ttk.Frame(self)
        toolbar.pack(fill="x", pady=(0, 8))
        ttk.Label(
            toolbar, text=window_title, font=("Segoe UI", 11, "bold")
        ).pack(side="left")
        ttk.Button(
            toolbar, text="Export to Excel", command=self._export_excel
        ).pack(side="right", padx=(6, 0))
        ttk.Button(
            toolbar, text="Print", command=self.print_sheet
        ).pack(side="right", padx=(6, 0))
        ttk.Button(
            toolbar, text="Open in Browser", command=self.open_in_browser
        ).pack(side="right")

        self.status_label = ttk.Label(
            self, text="Loading ledgers...", wraplength=1600,
        )
        self.status_label.pack(anchor="w", pady=(0, 8))

        # ── Compact treeview style ─────────────────────────────────────────
        style = ttk.Style()
        style.configure("Ledger.Treeview",         font=("Consolas", 8))
        style.configure("Ledger.Treeview.Heading", font=("Consolas", 8, "bold"))

        # ── Three side-by-side panels ─────────────────────────────────────
        paned = ttk.Panedwindow(self, orient="horizontal")
        paned.pack(fill="both", expand=True)

        self.panels: list[_LedgerPanelWidgets] = []
        for _ in range(3):
            panel = self._make_ledger_panel(paned)
            self.panels.append(panel)
            paned.add(panel.frame, weight=1)

        self.load_ledgers()

    # ── Panel factory ─────────────────────────────────────────────────────

    def _make_ledger_panel(self, parent) -> _LedgerPanelWidgets:
        frame = ttk.Frame(parent)

        title_label = ttk.Label(frame, font=("Segoe UI", 9, "bold"))
        title_label.pack(anchor="w")

        subtitle_label = ttk.Label(frame, font=("Segoe UI", 8))
        subtitle_label.pack(anchor="w", pady=(0, 4))

        table = ttk.Frame(frame)
        table.pack(fill="both", expand=True)
        table.rowconfigure(0, weight=1)
        table.columnconfigure(0, weight=1)

        col_ids = [col_id for col_id, _, _, _ in TREE_COLUMNS_COMPACT]
        tree = ttk.Treeview(
            table,
            columns=col_ids,
            show="headings",
            height=26,
            style="Ledger.Treeview",
        )
        tree.grid(row=0, column=0, sticky="nsew")

        for col_id, heading, width, anchor in TREE_COLUMNS_COMPACT:
            tree.heading(col_id, text=heading, anchor=anchor)
            stretch = col_id == "note"
            tree.column(col_id, width=width, minwidth=width, anchor=anchor, stretch=stretch)

        y_scroll = ttk.Scrollbar(table, orient="vertical",   command=tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(table, orient="horizontal", command=tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        return _LedgerPanelWidgets(frame, title_label, subtitle_label, tree)

    # ── Data loading ──────────────────────────────────────────────────────

    def _load_stock_ledger(
        self, stock_id: int, multiplier: float
    ) -> tuple[list[LedgerRow], int, tuple[str, str, str]]:
        prices = db.stock_price_list(stock_id)
        rows = build_ledger(prices, multiplier=multiplier)
        return rows, len(prices), threshold_labels(multiplier)

    def load_ledgers(self):
        group = self.group
        try:
            fresh = db.group_get(int(group["GroupID"])) or group

            stock_specs = [
                (int(fresh["Stock1ID"]), float(fresh.get("Exchange1Multiplier") or 1),
                 fresh["Stock1Ticker"]),
                (int(fresh["Stock2ID"]), float(fresh.get("Exchange2Multiplier") or 1),
                 fresh["Stock2Ticker"]),
                (int(fresh["Stock3ID"]), float(fresh.get("Exchange3Multiplier") or 1),
                 fresh["Stock3Ticker"]),
            ]

            self._stocks_data = []
            for stock_id, multiplier, ticker in stock_specs:
                rows, count, thresholds = self._load_stock_ledger(stock_id, multiplier)
                title = f"{ticker}"
                self._stocks_data.append((title, rows, count, thresholds))

        except Exception as exc:
            self.status_label.configure(text=str(exc))
            return

        for panel, (title, rows, count, thresholds) in zip(self.panels, self._stocks_data):
            self._populate_panel(panel, title, rows, count, thresholds)

        group_name  = group.get("GroupName", f"Group #{group['GroupID']}")
        tickers     = " / ".join(t for t, _, _, _ in self._stocks_data)
        group_title = f"{group_name}: {tickers}"

        html_stocks = [
            (title, rows, count, thresholds[0], thresholds[1])
            for title, rows, count, thresholds in self._stocks_data
        ]
        self._html = build_group_ledger_html(group_title, html_stocks)

        self.status_label.configure(
            text=(
                "Grid markers: * pivot  ^ near upper pivot (blue in print)  "
                "! near lower pivot (red in print).  "
                "Blank cells = no new extreme recorded that day."
            )
        )

    def _populate_panel(
        self,
        panel: _LedgerPanelWidgets,
        title: str,
        rows: list[LedgerRow],
        price_count: int,
        thresholds: tuple[str, str, str],
    ) -> None:
        panel.title.configure(text=title)
        if not rows:
            panel.subtitle.configure(
                text="No price history — pull prices for this stock first."
            )
        else:
            reversal_label, continuation_label, near_pivot_label = thresholds
            panel.subtitle.configure(
                text=(
                    f"{price_count} day(s) · "
                    f"{reversal_label} trend · {continuation_label} cont · "
                    f"{near_pivot_label} near-pivot · * pivot  ^ upper  ! lower"
                )
            )

        for item in panel.tree.get_children():
            panel.tree.delete(item)
        for row in rows:
            panel.tree.insert("", "end", values=_row_values(row))

    # ── Exports ───────────────────────────────────────────────────────────

    def _write_html_tempfile(self) -> Path:
        gid = self.group.get("GroupID", "group")
        temp = Path(tempfile.gettempdir()) / f"stock_group_{gid}_ledgers.html"
        temp.write_text(self._html, encoding="utf-8")
        return temp

    def open_in_browser(self):
        if not self._html:
            messagebox.showwarning("Ledger", "No ledger data to display.")
            return
        webbrowser.open(self._write_html_tempfile().as_uri())

    def print_sheet(self):
        if not self._html:
            messagebox.showwarning("Ledger", "No ledger data to print.")
            return
        webbrowser.open(self._write_html_tempfile().as_uri())
        messagebox.showinfo(
            "Print",
            "The printable ledger sheet opened in your browser.\n"
            "Use Ctrl+P (or the browser Print menu) to print.",
        )

    def _export_excel(self):
        if not self._stocks_data:
            messagebox.showwarning("Export", "No ledger data to export.")
            return
        _export_to_excel(self.group, self._stocks_data)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def open_ledger_sheet(parent, group: dict):
    gid   = group.get("GroupID", "?")
    gname = group.get("GroupName", f"Group #{gid}")
    window = tk.Toplevel(parent)
    window.title(f"Livermore Ledgers — {gname}")
    window.geometry("1800x900")
    sheet = LedgerSheetWindow(window, group)
    sheet.pack(fill="both", expand=True)
